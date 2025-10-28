import openpyxl
import magic
from django.http import HttpResponse
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from apps.core.models import TestCaseMetric, TestCaseModel, Module, TestPlan, PriorityChoice, HistoryTestPlan, Project, TestPlanSession
from apps.core.utils import TestcaseImportExcel
from apps.core.apis.serializers import TestcaseListSerializer, FileUploadSerializer, \
    TestMetrixSerializer, ModuleSerializer, TestPlanSerializer, TestScoreSerializer, \
    TestCaseNameSerializer, CreateTestPlanSerializer, TestPlanningSerializer, PlanSerializer, TestCaseOptionSerializer, \
    TestCaseScoreSerializer, PlanHistorySerializer, MetrixSerializer, HistoryPlanDetailsSerializer, \
    TestplanSessionSerializer, SessionSerializer, TestCaseSerializer, SearchTestCaseSerializer
from apps.core.utils import QueryHelpers
from django.db.models import Max, IntegerField
from drf_spectacular.utils import extend_schema
from apps.core.pagination import CustomPagination
from apps.core.apis.serializers import AITestPlanSerializer
from sentriQA.helpers import custom_generics as c
from aimode.chatbot import get_llm_response
from django.db.models.functions import Coalesce
from apps.core.helpers import generate_score, generate_session_id
from django.contrib.postgres.search import SearchVector, SearchQuery
from sentriQA.helpers.renders import ResponseInfo
from apps.core.filters import TestcaseFilter


@extend_schema(tags=["Modules List API"])
class ModuleAPIView(generics.ListAPIView):

    serializer_class = ModuleSerializer
    queryset = Module.objects.all()


@extend_schema(tags=["Testcase List API"])
class TestCaseList(c.CustomListCreateAPIView):

    def get_queryset(self):
        queryset = (TestCaseModel.objects.select_related('module')
                    .values('id', 'name', 'module__name', 'testcase_type', 'priority', 'status'))
        return queryset

    pagination_class = CustomPagination
    filter_class = TestcaseFilter
    serializer_class = TestcaseListSerializer


@extend_schema(tags=["Testcase Create API"])
class TestCaseView(c.CustomCreateAPIView):

    serializer_class = TestCaseSerializer


@extend_schema(tags=["Testcase Detail API"])
class TestCaseDetail(c.CustomRetrieveUpdateDestroyAPIView):

    serializer_class = TestCaseSerializer

    def get_object(self):
        queryset = (TestCaseModel.objects.select_related('module', 'project')
                    .prefetch_related('metrics').get(pk=self.kwargs['pk']))
        return queryset


class SearchAPIView(generics.ListAPIView):

    def get_queryset(self):
        queryset = (TestCaseModel.objects.select_related('module', 'project')
        .annotate(
            search=SearchVector('id', 'name', 'module__name', 'testcase_type', 'priority', 'status')
        ).filter(
            search=SearchQuery(self.request.GET.get('q')),
        ))
        return queryset

    pagination_class = CustomPagination
    serializer_class = SearchTestCaseSerializer


@extend_schema(tags=["Testcase Excel Upload API"])
class FileUploadView(generics.GenericAPIView):

    serializer_class = FileUploadSerializer

    def post(self, request, *args, **kwargs):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            instance = TestcaseImportExcel(self.request.data['file_name'])
            upload_status = instance.import_data()
            if instance and upload_status:
                return ResponseInfo.success_response(data=upload_status, message="Upload Successful")
            else:
                return ResponseInfo.error_response(error="Error", message='Error While Saving Data', status_code=status.HTTP_400_BAD_REQUEST)
        return ResponseInfo.error_response(error=serializer.errors, status_code=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["Testcase Plan Creation API"])
class TestPlanningView(generics.GenericAPIView):

    serializer_class = TestPlanSerializer

    def post(self, request, *args, **kwargs):
        serializer = TestPlanSerializer(data=request.data)
        if serializer.is_valid():
            score = generate_score(request.data)
            if score:
                return ResponseInfo.success_response(data=score, message="Test Plan Creation Successful")
            else:
                return ResponseInfo.error_response(error="Error")
        return ResponseInfo.error_response(error=serializer.errors, status_code=status.HTTP_400_BAD_REQUEST)


class TestPlanVersionAPI(c.CustomCreateAPIView):

    serializer_class = TestplanSessionSerializer


class GetTestVersionAPI(c.CustomGenericAPIView):

    serializer_class = SessionSerializer

    def get_queryset(self):
        queryset = TestPlanSession.objects.only('id', 'version', 'status').filter(session__session_id=self.kwargs.get('token'))
        return queryset


class VersionDetailAPI(generics.GenericAPIView):

    serializer_class = TestplanSessionSerializer

    def get_queryset(self):
        queryset = TestPlanSession.objects.filter(session__session_id=self.kwargs.get('token'),
                                                  version=self.kwargs.get('version'))
        return queryset

    def get(self, requests, *args, **kwargs):
        try:
            response = self.get_serializer(self.get_queryset(), many=True)
            if response.data:
                if isinstance(response.data, list):
                    if len(response.data) == 1:
                        return ResponseInfo.success_response(data=response.data[0], message="Test Plan Version Creation Successful")
                return ResponseInfo.success_response(data=response.data, message="Test Plan Version Creation Successful")
            else:
                return ResponseInfo.success_response(data=response.data, status_code=status.HTTP_200_OK)
        except Exception as e:
            return ResponseInfo().error_response(error={"error": str(e)}, status_code=status.HTTP_400_BAD_REQUEST)


class CreateTestPlanView(generics.GenericAPIView):
    serializer_class = CreateTestPlanSerializer

    def post(self, request, *args, **kwargs):
        try:
            serializer = CreateTestPlanSerializer(data=request.data)
            if serializer.is_valid():
                data = serializer.save()
                if data:
                    queryset = TestPlan.objects.get(id=data)
                    serializer = PlanSerializer(queryset)
                    return ResponseInfo.success_response(data=serializer.data, message="Test Plan Creation Successful")
                return ResponseInfo.error_response(error=serializer.errors, status_code=status.HTTP_400_BAD_REQUEST)
            return ResponseInfo.error_response(error={f"error: {serializer.errors}"}, status_code=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return ResponseInfo.error_response(error={f"error: {str(e)}"}, status_code=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["Classic Options API"])
class ClassicOptionAPI(APIView):

    def get(self, request, *args, **kwargs):
        # get_functionality = TestCaseModel.objects.values('testcase_type').distinct('testcase_type')
        get_project = Project.objects.all().values('id', 'name')
        get_priority = [{"id": value, "name": label} for value, label in PriorityChoice.choices]
        response_format = {
            "status": status.HTTP_200_OK,
            "data": {
                'testcase_type': [{
                    "id": "functional",
                    "name": "Functional",
                }],
                'priority': get_priority,
                'project': get_project
            },
            "status_code": status.HTTP_400_BAD_REQUEST,
            "message": "Success",
        }
        return Response(response_format, status=status.HTTP_200_OK)


class TestcaseOptionAPI(generics.GenericAPIView):

    def __init__(self, **kwargs):
        self.response_format = {
            "status": True,
            "data": {},
            "message": "Success",
            "status_code": status.HTTP_200_OK
        }
        super().__init__(**kwargs)

    serializer_class = MetrixSerializer

    def get_queryset(self):
        q = self.request.query_params.get('search', None)
        query = TestCaseMetric.objects.select_related(
            "testcase", "testcase__module"
        ).exclude(
            testcase__id__in=self.request.data.get('testcases', [])
        )
        query = query.annotate(
            max_time=Coalesce(
                Max('execution_time', output_field=IntegerField()),
                0
            )
        )
        if q.isdigit():
            queryset = query.filter(testcase__id=int(q))
            return queryset
        queryset = query.annotate(
            search=SearchVector("testcase__name", "testcase__module__name",
                                "testcase__testcase_type", "testcase__priority")
        ).filter(
            search=SearchQuery(q)
        ).distinct()
        return queryset

    def post(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        if serializer.data:
            return ResponseInfo.success_response(data=serializer.data, message="Test Plan Creation Successful")
        return ResponseInfo.error_response(error='Error', status_code=status.HTTP_400_BAD_REQUEST)


class TestcaseScore(generics.GenericAPIView):

    serializer_class = TestCaseScoreSerializer

    def get_queryset(self):
        testcase_metrix = TestCaseMetric.objects.filter(testcase__id__in=self.kwargs['pk'])
        return testcase_metrix

    def post(self, request, *args, **kwargs):
        serializer = TestMetrixSerializer(data=self.queryset(), many=True)
        if serializer.data:
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["AI Testcase Plan Creation API"])
class AITestPlanningView(generics.GenericAPIView):
    serializer_class = AITestPlanSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            user_msg = serializer.validated_data['user_msg']
            session = serializer.validated_data.get('session_id')
            if not session or session == "":
                session = generate_session_id()
            response_dict = get_llm_response(user_msg, session)
            response_dict['session_id'] = session
            if response_dict['tcs_data']:
                response_dict['chat_generated'] = True
            return Response(response_dict, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


@extend_schema(tags=["Testcase Plan List API"])
class TestPlanView(APIView):

    serializer_class = TestMetrixSerializer

    def post(self, request, *args, **kwargs):
        modules = request.data.get('module')

        if not modules:
            return Response(
                {'error': 'module parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Convert single module to list
        if not isinstance(modules, list):
            modules = [modules]

        queryset = TestCaseModel.objects.filter(module__name__in=modules)
        sub_query = TestCaseMetric.objects.filter(testcase__in=queryset)
        serializer = TestMetrixSerializer(sub_query, many=True)
        return Response(serializer.data)


@extend_schema(tags=["Testcase Score List API"])
class TestScores(APIView):

    def get(self, request, *args, **kwargs):
        queryset = TestCaseMetric.objects.all()
        serializer = TestMetrixSerializer(queryset, many=True)
        return Response(serializer.data)

@extend_schema(tags=["GetExcel API"])
class TestScoreExcel(APIView):

    def get(self, request, *args, **kwargs):
        wb = openpyxl.load_workbook('templates/Book_Test.xlsx')
        sheet = wb['Sheet1']
        row_num = 2
        for row in sheet.iter_rows(
            min_row=2, values_only=True,
        ):
            testcase_instance = QueryHelpers.get_test_case_instance(
                value=row[1]
            )
            try:
                _instance = TestCaseMetric.objects.get(testcase=testcase_instance)
                sheet.cell(row_num, column=17).value = round(_instance.get_test_scores())
                row_num += 1
            except Exception as e:
                print(e)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="test_score.xlsx"'
        wb.save(response)
        return response


class GetScoreViewAPIView(generics.GenericAPIView):
    serializer_class = TestCaseNameSerializer

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            queryset = TestCaseMetric.objects.filter(testcase__name__in=request.data['testcase'])
            serializer = TestScoreSerializer(data=request.data)
            if serializer.is_valid():
                return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TestPlanView(generics.GenericAPIView):

    serializer_class = TestPlanningSerializer
    queryset = TestPlan.objects.all()

    def get(self, request, *args, **kwargs):
        serializer = self.get_serializer(
            self.get_queryset(),
            many=True
        )
        if serializer.data:
            response_format = {
                "status": status.HTTP_200_OK,
                "data": serializer.data,
                "status_code": status.HTTP_200_OK,
                "message": "success",
            }
            return Response(response_format, status=status.HTTP_200_OK)
        else:
            response_format = {
                "status": status.HTTP_400_BAD_REQUEST,
                "data": None,
                "status_code": status.HTTP_400_BAD_REQUEST,
                "message": serializer.errors if serializer.errors else 'No data Found',
            }
            return Response(response_format, status=status.HTTP_400_BAD_REQUEST)


class PlanDetailsView(generics.RetrieveUpdateDestroyAPIView):

    serializer_class = PlanSerializer

    def get_object(self):
        queryset = TestPlan.objects.get(id=self.kwargs['id'])
        print(queryset)
        return queryset
    
    def put(self, request, *args, **kwargs):
        serializer = self.get_serializer(instance=self.get_object(), data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ConvertAPIView(APIView):

    def get_modules(self, modules):
        if modules:
            modules = Module.objects.filter(name__in=modules).values('id', 'name')
            return list(modules)
        return []

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            modules = request.data.get('module') if request.data.get('module') else []
            name = request.data.get('name') if 'name' in request.data else ""
            description = request.data.get('description') if 'description' in request.data else ""
            priority = request.data.get('priority') if 'priority' in request.data else ""
            response_format = {
                "status": "Success",
                "data": {
                    "name": name,
                    "description": description,
                    "priority": priority,
                    "modules": self.get_modules(modules),
                },
                "status_code": status.HTTP_200_OK,
                "message": "Success"
            }
            return Response(
                response_format,
                status=status.HTTP_200_OK
            )
        return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "data": None,
                "status_code": status.HTTP_400_BAD_REQUEST,
                "message": "ERROR"
            }, status=status.HTTP_400_BAD_REQUEST)
    

class TestPlanHistoryView(c.CustomListCreateAPIView):

    serializer_class = PlanHistorySerializer
    pagination_class = CustomPagination

    def get_queryset(self):
        testplan_id = self.kwargs['id']
        queryset = HistoryTestPlan.objects.filter(testplan_id=testplan_id).order_by('-created')
        return queryset
        

class HistoryPlanDetailsView(c.CustomRetrieveAPIVIew):

    serializer_class = HistoryPlanDetailsSerializer

    def get_object(self):
        queryset = HistoryTestPlan.objects.get(id=self.kwargs['history_id'])
        return queryset