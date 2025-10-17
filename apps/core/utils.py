from abc import ABC, abstractmethod
import pandas as pd
from django.db import transaction
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from openpyxl import load_workbook
from apps.core.helpers import QueryHelpers, generate_score
from apps.core.models import TestCaseModel, TestCaseMetric


class FileFactory(ABC):

    @abstractmethod
    def import_data(self):
        pass


class ExcelFileFactory(FileFactory):
    """Base factory class to handle Excel file operations."""

    def __init__(self, file):
        self.response_format = {
            "status": True,
            "status_code": HTTP_200_OK,
            "data": "",
            "message": ""
        }
        self.file = file

    def _init_workbook(self):
        """Initialize the workbook and return the active sheet."""
        workbook = load_workbook(self.file)
        return workbook.active

    def import_data(self):
        pass


class DemoExcelFileFactory(ExcelFileFactory):

    def __init__(self, file):
        super().__init__(file)
        self.ws = self._init_workbook()

    def get_failure_rate(self, fail, runs, weight):
        failure_rate = fail / runs
        return failure_rate * weight

    def get_impact_value(self, value):
        if value == 'Yes':
            return 1
        return 0.5

    def import_data(self):
        try:
            lst = []
            matrix = []
            for row in self.ws.iter_rows(min_row=3, values_only=True):
                if row[1] is not None:
                    temp = {
                        "name": row[1]
                    }
                    lst.append(TestCaseModel(**temp))
                    _matrix = {
                        "testcase": row[1],
                        "likelihood": row[4],
                        "impact": row[5],
                        "failure_rate": row[7],
                        "failure": row[8],
                        "total_runs": row[9],
                        "direct_impact": self.get_impact_value(row[10]),
                        "defects": row[12],
                        "severity": row[13],
                        "feature_size": row[14],
                        "execution_time": row[15]
                    }
                    print(_matrix)
                    matrix.append(_matrix)
            TestCaseModel.objects.bulk_create(lst)
            for i in range(len(matrix)):
                matrix[i]['testcase'] = QueryHelpers.get_test_case_instance(matrix[i]['testcase'])
                matrix[i] = TestCaseMetric(**matrix[i])
            with transaction.atomic():
                TestCaseMetric.objects.bulk_create(matrix)
            return True
        except Exception as e:
            print('error', e)
            return False


class TestcaseImportExcel(ExcelFileFactory):

    def __init__(self, file):
        super().__init__(file)
        self.ws = self._init_workbook()

    def _build_error_response(self, error):
        self.response_format.update({
            "status": False,
            "status_code": HTTP_400_BAD_REQUEST,
            "message": str(error),
        })
        return self.response_format

    def _build_success_response(self, message):
        self.response_format.update({
            "data": "Success",
            "message": message,
        })
        return self.response_format

    @staticmethod
    def get_priority(name):
        match name:
            case 'Class 1':
                return 'class_1'
            case 'Class 2':
                return 'class_2'
            case 'Class 3':
                return 'class_3'
        return False


    @staticmethod
    def get_failure_rate(failure, total_runs):
        if failure and total_runs:
            return (failure / total_runs) * 100
        else:
            return 0
    

    @staticmethod
    def get_impact_value(value):
        if value == 'Yes' or 'yes' or value == 0:
            return 1.0
        else:
            return 0.5


    def import_data(self):
        # try:
            current_module = None
            lst = []
            matrix = []
            for row in self.ws.iter_rows(
                min_row=2, values_only=True
            ):
                project_inst = QueryHelpers.get_project_by_id(name='nature')
                current_module = QueryHelpers.get_module_instance(row[1])
                if row[1]:
                    if not QueryHelpers.check_testcase_exists(row[4]):
                        _temp = {
                            "name": row[4] if row[4] is not None else "",
                            "priority": self.get_priority(row[7]),
                            "status": "completed",
                            "module": current_module,
                            "project": project_inst
                        }
                        lst.append(TestCaseModel(**_temp))
                if row[1] is not None:
                    _matrix = {
                        "testcase": row[4],
                        "likelihood": row[5],
                        "impact": row[6],
                        "failure_rate": self.get_failure_rate(row[10], row[11]),
                        "failure": row[10],
                        "total_runs": row[11],
                        "direct_impact": self.get_impact_value(row[12]),
                        "defects": row[14],
                        "severity": 0,
                        "feature_size": 0,
                        "execution_time": 0
                    }
                    matrix.append(_matrix)
            if lst:
                TestCaseModel.objects.bulk_create(lst)
                
            for i in range(len(matrix) - 1, -1, -1):
                if not QueryHelpers.check_matrix_id(matrix[i]['testcase']):
                    matrix[i]['testcase'] = QueryHelpers.get_test_case_instance(matrix[i]['testcase'])
                    matrix[i] = TestCaseMetric(**matrix[i])
            with transaction.atomic():
                TestCaseMetric.objects.bulk_create(matrix)
            return True
        # except Exception as e:
        #     print('error', e)
        #     return False
